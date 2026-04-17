"""
Generador de Excel — Cuadros de Carga Alumbrado Público
openpyxl — hoja RESUMEN + SUPUESTOS + hoja por empalme con detalle poste a poste
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paleta ─────────────────────────────────────────────────────────────────────
AZUL_H   = "1F3864"
AZUL_CL  = "D6E4F0"
GRIS_F   = "F2F2F2"
VERDE_OK = "C6EFCE"
ROJO_NOK = "FFC7CE"
AMARILLO = "FFEB9C"
BLANCO   = "FFFFFF"
NEGRO    = "000000"
COL_R    = "FFD7D7"
COL_S    = "D7FFD7"
COL_T    = "D7E8FF"

FASE_BG = {"R": COL_R, "S": COL_S, "T": COL_T}


def _font(bold=False, size=9, color=NEGRO):
    return Font(name="Arial", bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _c(ws, row, col, value, bold=False, bg=None, fg=NEGRO, h="center", size=9):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=size, color=fg)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align(h=h)
    cell.border = _border()
    return cell

def _hdr(ws, row, headers, col_start=1, bg=AZUL_H, fg=BLANCO, h=20):
    for i, hdr in enumerate(headers, start=col_start):
        _c(ws, row, i, hdr, bold=True, bg=bg, fg=fg)
    ws.row_dimensions[row].height = h

def _cw(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Hoja RESUMEN ───────────────────────────────────────────────────────────────
def _hoja_resumen(wb, empalmes, proyecto):
    ws = wb.active
    ws.title = "RESUMEN"
    from openpyxl.utils import get_column_letter

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"RESUMEN GENERAL — {proyecto.get('nombre','CUADROS DE CARGA')}"
    c.font = _font(bold=True, size=12, color=BLANCO)
    c.fill = _fill(AZUL_H)
    c.alignment = _align()
    ws.row_dimensions[1].height = 24

    # Meta
    meta = [
        ("Proyecto:", proyecto.get("nombre","—")),
        ("Tramo:", proyecto.get("tramo","—")),
        ("DM Inicial:", proyecto.get("dm_ini","—")),
        ("DM Final:", proyecto.get("dm_fin","—")),
        ("Ingeniero:", proyecto.get("ingeniero","—")),
        ("Fecha:", proyecto.get("fecha", str(date.today()))),
        ("Normativa:", "SEC RIC / Manual Carreteras MOP"),
    ]
    row = 2
    for i, (k, v) in enumerate(meta):
        ck = 1 + (i % 2) * 2
        cv = ck + 1
        if i % 2 == 0 and i > 0:
            row += 1
        _c(ws, row, ck, k, bold=True, bg=AZUL_CL, h="right", size=8)
        _c(ws, row, cv, v, bg=AZUL_CL, h="left", size=8)
        ws.row_dimensions[row].height = 14
    row += 2

    # Título sección
    ws.merge_cells(f"A{row}:H{row}")
    _c(ws, row, 1, "TABLA RESUMEN POR EMPALME", bold=True, bg=AZUL_H, fg=BLANCO, h="left", size=9)
    ws.row_dimensions[row].height = 16
    row += 1

    headers = ["Empalme", "Tipo", "N° Circ.", "N° Lum.",
               "P. Inst.(kW)", "P. Máx.(kW)", "I Máx.(A)", "Disyuntor G.(A)"]
    _hdr(ws, row, headers)
    row += 1

    pot_t = lum_t = 0
    for emp in empalmes:
        bg = BLANCO if row % 2 == 1 else GRIS_F
        for col, val in enumerate([
            emp["id"], emp["tipo"], emp["n_circuitos"], emp["n_luminarias_total"],
            emp["pot_instalada_kw"], emp["pot_maxima_kw"],
            emp["corriente_max_a"], f"{emp['disyuntor_gral_a']}A",
        ], 1):
            _c(ws, row, col, val, bg=bg, size=8)
        ws.row_dimensions[row].height = 14
        pot_t += emp["pot_instalada_kw"]
        lum_t += emp["n_luminarias_total"]
        row += 1

    for col, val in enumerate([
        "TOTAL PROYECTO", "—",
        sum(e["n_circuitos"] for e in empalmes),
        lum_t, round(pot_t, 3), round(pot_t, 3), "—", "—",
    ], 1):
        _c(ws, row, col, val, bold=True, bg=AZUL_H, fg=BLANCO, size=9)
    ws.row_dimensions[row].height = 16

    _cw(ws, {"A":14,"B":12,"C":12,"D":14,"E":16,"F":16,"G":14,"H":18})


# ── Hoja por empalme ───────────────────────────────────────────────────────────
def _hoja_empalme(wb, emp, proyecto):
    ws = wb.create_sheet(title=emp["id"][:31])

    # Título
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = (f"CUADRO DE CARGAS — {emp['id']} ({emp['tipo']}) — "
               f"{proyecto.get('nombre','')}")
    c.font = _font(bold=True, size=11, color=BLANCO)
    c.fill = _fill(AZUL_H)
    c.alignment = _align()
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = (f"Tramo: {proyecto.get('tramo','')}  |  "
               f"DM {proyecto.get('dm_ini','')}–{proyecto.get('dm_fin','')}  |  "
               f"Ing.: {proyecto.get('ingeniero','')}  |  "
               f"Fecha: {proyecto.get('fecha', str(date.today()))}")
    c.font = _font(size=8)
    c.fill = _fill(AZUL_CL)
    c.alignment = _align()
    ws.row_dimensions[2].height = 14

    row = 3

    # ── 1. Resumen de circuitos ────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:N{row}")
    _c(ws, row, 1, "1. RESUMEN DE CIRCUITOS", bold=True, bg=AZUL_H, fg=BLANCO, h="left", size=9)
    ws.row_dimensions[row].height = 16
    row += 1

    hdrs_cto = ["Cto.", "Fases", "N° Lum.", "P.Inst.(kW)",
                "I calc.(A)", "I dis.(A)", "Disyuntor(A)",
                "Secc.(mm²)", "Long.(m)", "ΔV máx.(%)", "Cumple ΔV",
                "Cumple Cond.", "Material", "Estado"]
    _hdr(ws, row, hdrs_cto)
    row += 1

    for c_data in emp["circuitos"]:
        ok_dv  = c_data["cumple_dv"]
        ok_cnd = c_data["cumple_conductor"]
        ok_all = ok_dv and ok_cnd
        bg = BLANCO if row % 2 == 1 else GRIS_F
        fases_str = "/".join(sorted(set(
            p["fase"] for p in c_data.get("postes",[]) if p["fase"] != "—"
        ))) or ("R/S/T" if emp["tipo"] == "3F" else "—")
        vals = [
            c_data["circuito"], fases_str, c_data["n_luminarias"],
            c_data["pot_instalada_kw"], c_data["corriente_calc_a"],
            c_data["corriente_diseno_a"], c_data["disyuntor_a"],
            f"{c_data['seccion_mm2']:.1f}", c_data["longitud_m"],
            f"{c_data['dv_pct']:.2f}%",
            "SÍ" if ok_dv  else "NO",
            "SÍ" if ok_cnd else "NO",
            c_data["material"],
            "CUMPLE" if ok_all else "REVISAR",
        ]
        for col, val in enumerate(vals, 1):
            bg_cell = bg
            if col == 11:
                bg_cell = VERDE_OK if ok_dv  else ROJO_NOK
            elif col == 12:
                bg_cell = VERDE_OK if ok_cnd else ROJO_NOK
            elif col == 14:
                bg_cell = VERDE_OK if ok_all else ROJO_NOK
            _c(ws, row, col, val, bg=bg_cell, size=8)
        ws.row_dimensions[row].height = 14
        row += 1

    # Totales
    for col, val in enumerate([
        "TOTAL", "—", emp["n_luminarias_total"],
        round(emp["pot_instalada_kw"], 3), "—", "—",
        f"{emp['disyuntor_gral_a']}A", "—", "—", "—", "—", "—", "—", "—",
    ], 1):
        _c(ws, row, col, val, bold=True, bg=AZUL_CL, size=8)
    ws.row_dimensions[row].height = 14
    row += 2

    # ── 2. Selección de conductor ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:P{row}")
    _c(ws, row, 1, "2. SELECCIÓN DE CONDUCTOR — TRES CRITERIOS NORMATIVOS (IEC 60364-4-43)",
       bold=True, bg=AZUL_H, fg=BLANCO, h="left", size=9)
    ws.row_dimensions[row].height = 16
    row += 1

    hdrs_cond = [
        "Cto.", "Mat.", "Aisl.", "k (A·s⁰·⁵/mm²)",
        "S corr.(mm²)", "S ΔV(mm²)", "S mín.ΔV calc.", "S Icc(mm²)", "S mín.Icc calc.",
        "→ S final(mm²)", "Criterio limitante",
        "Ampacidad(A)", "Icc_adm.(A)", "ΔV verif.(%)",
        "✔ Corriente", "✔ ΔV", "✔ Icc",
    ]
    _hdr(ws, row, hdrs_cond, h=20)
    row += 1

    for c_data in emp["circuitos"]:
        cs = c_data.get("conductor_seleccion", {})
        if not cs:
            continue
        ok_i = cs.get("cumple_corriente", True)
        ok_d = cs.get("cumple_dv", True)
        ok_c = cs.get("cumple_cc", True)
        bg = BLANCO if row % 2 == 1 else GRIS_F
        vals = [
            c_data["circuito"],
            c_data["material"],
            c_data.get("tipo_aislamiento", "XLPE"),
            cs.get("k_adiab", "—"),
            cs.get("s_por_corriente_mm2", "—"),
            cs.get("s_por_dv_mm2", "—"),
            f"{cs.get('s_dv_min_calc',0):.3f}",
            cs.get("s_por_cc_mm2", "—"),
            f"{cs.get('s_cc_min_calc',0):.3f}",
            cs.get("seccion_mm2", "—"),
            cs.get("criterio_limitante", "—"),
            cs.get("capacidad_a", "—"),
            f"{cs.get('i_cc_admisible_a',0):.1f}",
            f"{cs.get('dv_verificacion_pct',0):.3f}%",
            "SÍ" if ok_i else "NO",
            "SÍ" if ok_d else "NO",
            "SÍ" if ok_c else "NO",
        ]
        for col, val in enumerate(vals, 1):
            if col == 10:
                bg_c = AZUL_CL
            elif col == 15:
                bg_c = VERDE_OK if ok_i else ROJO_NOK
            elif col == 16:
                bg_c = VERDE_OK if ok_d else ROJO_NOK
            elif col == 17:
                bg_c = VERDE_OK if ok_c else ROJO_NOK
            else:
                bg_c = bg
            bold_c = col == 10
            _c(ws, row, col, val, bold=bold_c, bg=bg_c, size=8)
        ws.row_dimensions[row].height = 14
        row += 1
    row += 1

    # ── 3. Detalle poste a poste ───────────────────────────────────────────────
    ws.merge_cells(f"A{row}:N{row}")
    _c(ws, row, 1, "3. DETALLE POSTE A POSTE POR CIRCUITO", bold=True,
       bg=AZUL_H, fg=BLANCO, h="left", size=9)
    ws.row_dimensions[row].height = 16
    row += 1

    hdrs_poste = ["Código", "Cto.", "Interdist.(m)", "Potencia(W)", "Fase",
                  "I_poste(A)", "I_seg(A)", "ΔV_tramo(V)", "ΔV_acum(V)", "ΔV_acum(%)"]

    for c_data in emp["circuitos"]:
        if not c_data.get("postes"):
            continue

        # Sub-encabezado de circuito
        ws.merge_cells(f"A{row}:J{row}")
        _c(ws, row, 1,
           f"Circuito {c_data['circuito']:02d} — "
           f"{c_data['n_luminarias']} lum. — "
           f"{c_data['pot_instalada_kw']:.3f} kW — "
           f"I={c_data['corriente_calc_a']:.2f}A — "
           f"Cond: {c_data['seccion_mm2']:.1f}mm² {c_data['material']}",
           bold=True, bg=AZUL_CL, h="left", size=8)
        ws.row_dimensions[row].height = 14
        row += 1

        _hdr(ws, row, hdrs_poste, h=18)
        row += 1

        for p in c_data["postes"]:
            bg_fase = FASE_BG.get(p["fase"], BLANCO)
            dv_ok   = p["dv_acum_pct"] <= 3.0
            bg_row  = BLANCO if row % 2 == 1 else GRIS_F

            vals = [
                p["codigo"],
                c_data["circuito"],
                p["interdistancia_m"],
                p["pot_w"],
                p["fase"],
                p["corriente_a"],
                p["i_segmento_a"],
                p["dv_tramo_v"],
                p["dv_acum_v"],
                p["dv_acum_pct"],
            ]
            for col, val in enumerate(vals, 1):
                if col == 5:
                    bg_cell = bg_fase
                elif col == 10:
                    bg_cell = VERDE_OK if dv_ok else ROJO_NOK
                else:
                    bg_cell = bg_row
                _c(ws, row, col, val, bg=bg_cell, size=7.5)
            ws.row_dimensions[row].height = 13
            row += 1

        # Fila max/total del circuito
        for col, val in enumerate([
            "MÁX/TOTAL", str(c_data["circuito"]),
            round(sum(p["interdistancia_m"] for p in c_data["postes"]), 1),
            round(sum(p["pot_w"]           for p in c_data["postes"]), 1),
            "—", "—", "—",
            "—",
            round(max(p["dv_acum_v"]   for p in c_data["postes"]), 4),
            round(max(p["dv_acum_pct"] for p in c_data["postes"]), 3),
        ], 1):
            bg_cell = AZUL_CL
            if col == 10:
                bg_cell = VERDE_OK if c_data["cumple_dv"] else ROJO_NOK
            _c(ws, row, col, val, bold=True, bg=bg_cell, size=8)
        ws.row_dimensions[row].height = 14
        row += 2

    # ── 3. Balance de fases (3F) ───────────────────────────────────────────────
    if emp["tipo"] == "3F" and emp["balance_fases"]:
        ws.merge_cells(f"A{row}:E{row}")
        _c(ws, row, 1, "4. BALANCE DE FASES", bold=True,
           bg=AZUL_H, fg=BLANCO, h="left", size=9)
        ws.row_dimensions[row].height = 16
        row += 1

        _hdr(ws, row, ["Fase", "Potencia(kW)", "Corriente(A)", "% Desbalance", "Estado"])
        row += 1

        bal = emp["balance_fases"]
        for f_name in ["R", "S", "T"]:
            fd = bal["fases"][f_name]
            ok = fd["desbalance_pct"] <= 10
            bg = BLANCO if row % 2 == 1 else GRIS_F
            vals = [f_name, fd["potencia_kw"], fd["corriente_a"],
                    f"{fd['desbalance_pct']:.2f}%", "CUMPLE" if ok else "REVISAR"]
            for col, val in enumerate(vals, 1):
                bg_c = VERDE_OK if (col == 5 and ok) else (ROJO_NOK if col == 5 else bg)
                _c(ws, row, col, val, bg=bg_c, size=8)
            ws.row_dimensions[row].height = 14
            row += 1

        for col, val in enumerate([
            "PROMEDIO", "—", f"{bal['i_promedio_a']:.2f}",
            f"Máx: {bal['desbalance_max_pct']:.2f}%",
            "CUMPLE" if bal["cumple"] else "REVISAR",
        ], 1):
            bg_c = VERDE_OK if (col == 5 and bal["cumple"]) else \
                   (ROJO_NOK if col == 5 else AZUL_CL)
            _c(ws, row, col, val, bold=True, bg=bg_c, size=8)
        ws.row_dimensions[row].height = 14
        row += 2

    # ── 4. Dimerización ───────────────────────────────────────────────────────
    sec = "5" if emp["tipo"] == "3F" else "4"
    ws.merge_cells(f"A{row}:D{row}")
    _c(ws, row, 1, f"{sec}. ESCENARIOS DE DIMERIZACIÓN", bold=True,
       bg=AZUL_H, fg=BLANCO, h="left", size=9)
    ws.row_dimensions[row].height = 16
    row += 1

    _hdr(ws, row, ["Escenario", "% Potencia", "Potencia(kW)", "Corriente(A)"])
    row += 1

    dim_bgs = ["D6E8FF", GRIS_F, AMARILLO]
    for i, esc in enumerate(emp["dimerizacion"]):
        bg = dim_bgs[i % 3]
        for col, val in enumerate([
            esc["escenario"], f"{esc['factor_pct']}%",
            esc["potencia_kw"], esc["corriente_a"]
        ], 1):
            _c(ws, row, col, val, bg=bg, size=8)
        ws.row_dimensions[row].height = 14
        row += 1

    # Ancho de columnas
    _cw(ws, {
        "A": 16, "B": 8,  "C": 9,  "D": 12, "E": 10,
        "F": 14, "G": 14, "H": 14, "I": 14, "J": 14,
        "K": 22, "L": 13, "M": 14, "N": 12,
        "O": 13, "P": 10, "Q": 10,
    })


# ── Hoja SUPUESTOS ─────────────────────────────────────────────────────────────
def _hoja_supuestos(wb):
    ws = wb.create_sheet(title="SUPUESTOS")

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "SUPUESTOS TÉCNICOS Y CRITERIOS NORMATIVOS"
    c.font = _font(bold=True, size=11, color=BLANCO)
    c.fill = _fill(AZUL_H)
    c.alignment = _align()
    ws.row_dimensions[1].height = 22

    _hdr(ws, 2, ["Parámetro", "Valor", "Fuente / Normativa"], h=22)

    supuestos = [
        ("Factor de simultaneidad",    "1.0",             "SEC — Alumbrado público"),
        ("Factor de potencia (FP)",    "0.95",            "Luminarias LED con driver electrónico"),
        ("Factor diseño protecciones", "1.25",            "SEC RIC — Dimensionamiento protecciones"),
        ("ΔV% máximo recomendado",     "3%",              "Manual Carreteras MOP — Alumbrado vial"),
        ("Desbalance máximo fases",    "10%",             "Buenas prácticas ingeniería eléctrica"),
        ("Tensión nominal 1F",         "220 V",           "SEC Chile — Sistema BT monofásico"),
        ("Tensión nominal 3F",         "380/220 V",       "SEC Chile — Sistema BT trifásico"),
        ("Conductividad Al",           "38 m/(Ω·mm²)",    "Conductores CALPE-AL / XLPE-AL"),
        ("Conductividad Cu",           "56 m/(Ω·mm²)",    "Conductores XLPE-CU"),
        ("ΔV 1F (ida+vuelta)",         "2×Σ(d_j×I_j)/(σ×S)", "Carga distribuida variable"),
        ("ΔV 3F (solo fase)",          "Σ(d_j×I_fase_j)/(σ×S)", "Fases alternas R/S/T por poste"),
        ("Corriente diseño",           "I_d = I_calc×1.25", "SEC RIC"),
        ("Código luminaria",           "EE.CC.NN",        "EE=empalme, CC=circuito, NN=luminaria"),
        ("Fases alternas",             "R→S→T→R→S→T...", "Balanceo automático por posición"),
        ("Punta 100%",                 "Factor 1.00",     "Todas las luminarias al máximo"),
        ("Media noche 70%",            "Factor 0.70",     "Dimerización — reducción energética"),
        ("Valle 50%",                  "Factor 0.50",     "Dimerización — madrugada bajo tráfico"),
    ]

    for i, (param, valor, fuente) in enumerate(supuestos, start=3):
        bg = BLANCO if i % 2 == 1 else GRIS_F
        _c(ws, i, 1, param, bold=True, bg=bg, h="left", size=8)
        _c(ws, i, 2, valor, bg=bg, size=8)
        _c(ws, i, 3, fuente, bg=bg, h="left", size=8)
        ws.row_dimensions[i].height = 14

    _cw(ws, {"A": 30, "B": 28, "C": 60})


# ── Función principal ──────────────────────────────────────────────────────────
def generar_excel(proyecto: dict, empalmes: list) -> bytes:
    wb = Workbook()
    _hoja_resumen(wb, empalmes, proyecto)
    _hoja_supuestos(wb)
    for emp in empalmes:
        _hoja_empalme(wb, emp, proyecto)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
